import openpyxl
from openpyxl.styles import NamedStyle, Border, Side
from typing import List
import os


class Excel:
    def __init__(self, file_path: str) -> None:
        self.excel_file = openpyxl.load_workbook(file_path)
        self.sheet = self.excel_file.active

    def get_data(self) -> List[List[str]]:
        rows = self.sheet.max_row
        columns = self.sheet.max_column
        res = []
        for row in range(1, rows + 1):
            row_data = []
            for col in range(1, columns + 1):
                cell = self.sheet.cell(row, col)
                cell_value = cell.value
                if col == 1 and type(cell_value) == str:
                    break
                if col == 2:
                    cell_value = cell_value.encode("utf-8")
                if cell_value == None:
                    cell_value = ""
                row_data.append(cell_value if type(cell_value)
                                == bytes else str(cell_value))
            if len(row_data) > 0:
                res.append(row_data)
        return res

    @staticmethod
    def get_template(save_path: str, data: List[List[str]]) -> None:
        wb = openpyxl.Workbook()
        sheet = wb.active
        ns = NamedStyle(name='highlight')
        border = Side(style='thin', color='000000')
        ns.border = Border(left=border, top=border,
                           right=border, bottom=border)
        wb.add_named_style(ns)
        for row_index, row_data in enumerate(data):
            sheet.append(row_data)
            for cell_index, _ in enumerate(row_data):
                sheet.cell(row_index + 1, cell_index + 1).style = 'highlight'
        wb.save(save_path)
